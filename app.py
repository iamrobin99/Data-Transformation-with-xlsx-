from datetime import *
from datetime import timedelta
import pandas as pd
import smtplib
import numpy as np
from email.message import EmailMessage
import openpyxl 
from openpyxl.styles import PatternFill

#path for openpyxl
wb = openpyxl.load_workbook(r'C:\Users\User\Downloads\POFU -Candidate Status (For Robin).xlsx')
ws = wb['POFU']
# Colouring the sheet
red = PatternFill(patternType='solid', 
                           fgColor='FC2C03')
green = PatternFill(patternType='solid',fgColor='35FC03')

yellow = PatternFill(patternType='solid',fgColor='FCBA03')

#path for dataframe
df = pd.read_excel(r'C:\Users\User\Downloads\POFU -Candidate Status (For Robin).xlsx',sheet_name='POFU',engine='openpyxl')
#df_ = df.loc[df['Recruiter connect-1 Day 2-7 Date']!=0]

#df = df.loc[df['Resignation Acceptance (Assuming today is 25th July-22)'] != '']
df['Recruiter connect-1 Day 2-7 Date'] = df['Recruiter connect-1 Day 2-7 Date'].fillna(0)
#df = df.loc[df['Recruiter connect-1 Day 2-7 Date'] == 0]

df['Resignation Acceptance (Assuming today is 25th July-22)'] = df['Resignation Acceptance (Assuming today is 25th July-22)'].fillna(0)
df = df.loc[df['Resignation Acceptance (Assuming today is 25th July-22)'] != 0]

dat = np.datetime64('today')

index = df.index.values
resgn_date = df['Resignation Acceptance (Assuming today is 25th July-22)'].values
rec_con = df['Recruiter connect-1 Day 2-7 Date'].values
rec_con_ageing = df['Recruiter connect-1 Day 2-7 Ageing'].values
#email = df['SPOC E-Mail ID'].values
zipped = zip(index,resgn_date,rec_con,rec_con_ageing)


for (a,c,d,e) in zipped:
    c = np.datetime64(c).astype('datetime64[D]')
    #d = np.datetime64(d).astype('datetime64[D]')
    # add 2 for index correction
    a = a + 2
    delta = int(np.busday_count(c,dat))
    delta_con = int(np.busday_count(c,d))
    #delta = np.timedelta64(delta,"D")
    print(type(delta))
    
    print(delta)
#    days = delta.astype('timedelta64[D]')
    if delta > 6 and (d == 0):
        ws['S{}'.format(a)].fill = red
        ws['S{}'.format(a)] = delta
    elif d != 0 or delta <= 4:
        ws['S{}'.format(a)].fill = green
        if d == 0:
            ws['S{}'.format(a)] = delta
        else:
            ws['S{}'.format(a)] = delta_con
    elif delta > 4 and delta <= 6:
        ws['S{}'.format(a)].fill = yellow
        ws['S{}'.format(a)] = delta
wb.save(r'C:\Users\User\Downloads\POFU -Candidate Status (For Robin).xlsx')
